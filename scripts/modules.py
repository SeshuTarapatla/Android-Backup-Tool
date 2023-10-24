# importing libraries

from msvcrt import getch
from numpy import cumsum, diff, linspace
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from os import environ, getcwd, makedirs, path, remove, system
from pandas import DataFrame, concat, read_excel
from shutil import move, rmtree
from subprocess import run
from sys import exit
from tqdm import tqdm


# defined variables and lambda functions

ENC = 'utf-8'
exec = lambda cmd: run(cmd,capture_output=True)
makedir = lambda dir: makedirs(dir,exist_ok=True)



# defined classes and functions

def initialize():
    # function that initializes tool
    system('title Android Device Backup')
    print('Loading . . . ',end='',flush=True)
    ADB.add_path()
    ADB.start()
    system('cls')

def safe_exit():
    # function to exit tool properly
    ADB.stop()
    if path.isdir('cache'): rmtree('cache')
    print('\nPress any key to exit . . . ',end='',flush=True)
    getch()
    exit()

def parse_config(device,cfile='data\\config.txt'):
    config,flag = {},False
    config['device'] = device
    if path.isfile(cfile):
        # parsing values from config file
        with open(cfile,'r') as file:
            for line in file.readlines():
                prop,value = line.split('=')
                config[prop.strip()] = value.strip()
    else:
        # if no config exists, default values
        config['mode'] = 1
        config['output'] = path.join(getcwd(),device)
    
    # fault check
    if config['mode'] in ['0','1','2']: config['mode'] = int(config['mode'])
    else: config['mode'] = 1
    try:
        makedir(config['output'])
        if config['output'].endswith(device): 
            config['output'] = config['output'][:-len(device)]
        flag = True
    except:
        config['output'] = getcwd()
    print(f'Backup directory: "{config["output"]}" | Mode: {config["mode"]}')
    
    #saving config as latest
    with open(cfile,'w') as file:
        file.write(f'mode={config["mode"]}\n')
        file.write(f'output={config["output"] if flag else ""}')
    return config

def subtask(task=None,indent=1):
    # function to print subtasks
    if task is None:
        print('\t'*indent+'Done')
    else:
        print(f' > {task}'+'\t'*indent+':',end='',flush=True)

class ADB:
    # class for ADB functions
    
    def add_path(dir='adb'):
        # function that add ADB to path
        adb_dir = path.join(getcwd(),dir)
        if adb_dir not in environ['PATH']:
            environ['PATH'] += f'{adb_dir};'
    
    def start():
        # starts ADB server
        exec('adb start-server')
    
    def stop():
        # stops ADB server
        exec('adb kill-server')
    
    def push(src,dst):
        # pushes a file to android
        return exec(f'adb push "{src}" "{dst}"')
    
    def pull(src,dst):
        # pulls a file from android
        return exec(f'adb pull "{src}" "{dst}"')
    
    def remove(file):
        # deletes a file from android
        return exec(f'adb shell rm "{file}"')
    
    def get_device():
        # function the get connected device details
        resp = exec('adb devices').stdout.decode().splitlines()
        if len(resp) < 3:
            print(f'No device is connected. Please refer readme (4.2).')
            safe_exit()
        else:
            id = resp[1].split('\t')[0]
            device = exec(f'adb -s {id} shell getprop ro.product.product.model').stdout.decode()[:-2]
            print(f'Device connected: {device} | ID: {id}')
            return (device,id)


class categories:
    # class for category and file type functions
    def __init__(self):
        self.wb = load_workbook('data\\categories.xlsx')
        self.sheet = self.wb.active
        self.load_data()
        self.save_data()
    
    def load_data(self):
        # function the loads data from categories.xlsx
        self.data = {}
        self.alltypes = set()
        for c in range(1,self.sheet.max_column+1):
            category = self.sheet.cell(1,c).value
            types = set()
            for r in range(2,self.sheet.max_row+1):
                type = str(self.sheet.cell(r,c).value)
                types.add(type)
                self.sheet.cell(r,c).value = None
            if 'None' in types: types.remove('None')
            self.alltypes.update(types)
            self.data[category] = sorted(types)
        self.filetypes = self.alltypes - set(self.data['Exclude'])
        self.categories = sorted(self.data.keys())
        self.categories.remove('Exclude')
    
    def get_category(self,type):
        # function the returns the category of a type
        if type not in self.filetypes: return None
        for category in self.data.keys():
            if type in self.data[category]:
                return category
    
    def save_data(self):
        # function that saves the data to categories.xlsx
        headers = self.categories + ['Exclude']
        for i,category in enumerate(headers):
            self.sheet.cell(1,1+i).value = category
            self.sheet.cell(1,1+i).font = Font(bold=True)
            for j,type in enumerate(self.data[category]):
                self.sheet.cell(2+j,1+i).value = type
        self.wb.save('data\\categories.xlsx')


class filename:
    def __init__(self,file):
        file = path.splitext(file)
        self.file = file[0]
        self.type = file[1]
        self.suffix = 0
        self.set_suffix()
    
    def set_suffix(self):
        if self.file.endswith(')'):
            part = self.file[:-1].split('(')+['']
            if part[1].isdigit():
                self.suffix = int(part[1])
                self.file = part[0]
    
    def filename(self):
        if self.suffix == 0:
            return f'{self.file}{self.type}'
        else:
            return f'{self.file} ({self.suffix}){self.type}'